import xlrd
import os
__author__ = 'M'
# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook

def is_secured(country):
    with open("secured_countries.txt") as f:
        for line in f:
            if line.startswith(country):
                return True
    return False


def open_file(path, year , country , new_pop):
    if new_pop == "-":
        return
    full_path = os.getcwd()+"\\"+path.replace("/","\\")
    wb_re_read = load_workbook(filename=full_path)
    sheet = wb_re_read.get_sheet_by_name("ESTIMATES")

    book = xlrd.open_workbook(path)
    first_sheet = book.sheet_by_index(0)

    i=0
    country_found = False
    for row_value in first_sheet.col_values(2):
        i += 1
        #print(row_value)
        if row_value == country:
            #print("I find it:  "+str(i))
            country_found = True
            break

    if not country_found :
        print("data for this country is not available!")
        return
    j=0
    found = False
    for row_value in first_sheet.row_values(16):
        j += 1
        if row_value == year:
            #print("I find it:  "+str(j))
            found = True
            break
    if not found:
        print("data for this year is not available!")
        return

    sheet.cell(row=i,column=j).value = new_pop
    wb_re_read.save(filename=full_path)




year = input("Please enter year:")
country = input("Please enter country name:")
men_pop = input("Please enter men population:")
women_pop = input("Please enter women population:")

if is_secured(country):
        print("sorry... you can't change this country population")
else:
    open_file("Data/WPP2015_POP_F01_2_TOTAL_POPULATION_MALE.xlsx",year , country , men_pop)
    open_file("Data/WPP2015_POP_F01_3_TOTAL_POPULATION_FEMALE.xlsx",year , country , women_pop)
